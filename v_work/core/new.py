import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_wokrbook(filename):

    wb = xl.load_workbook(filename)

    sheets = wb.sheetnames
    print(sheets)

    ws = wb.active
    print(ws['A2'].value)

    #ws = wb.get_active_sheet()
    #print(ws['A1'].value)

    #cell = ws.cell(3, 4)
    #print(cell.value)

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = ws.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(ws,
              min_row=2,
              max_row=ws.max_row,
              min_col=4,
              max_col=4)

    chart = BarChart()
    chart.add_data(values)
    ws.add_chart(chart, 'e2')

    wb.save(filename)



